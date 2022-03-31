# excel2xml.py

from openpyxl import load_workbook
from yattag import Doc, indent
from datetime import date

po = str(input('Enter PO number: '))
lines = int(input('Enter last line number used in the file: '))

w = po+'.xlsx'
x = lines
date = date.today()

## Checks for Null or "" cells
## and replace all empty cells with the string None.
def clean_WB(w, x):
    wb = load_workbook(w)
    ws = wb.worksheets[0]
    for row in ws.iter_rows(min_row=2, max_row= x, min_col=1, max_col=14):
        for cell in row:
##            print(cell.value) ## Use to verify output when necessary
            if cell.value is None:
                cell.value = "None"
    wb.save('clean_wb.xlsx')
    

def convert2xml(w, x, date):
    # Loads Excel workbook 
    wb = load_workbook(w)

    # Create sheet object
    ws = wb.worksheets[0]
    wblist = []

    ### Returning returns a triplet
    doc, tag, text = Doc().tagtext()

    xml_header = '<?xml version= "1.0" encoding="UTF-8"?>'
    xml_schema = '<xs:schema xmlns:xs="http://www.w3.org/2001/XMLSchema"></xs:schema>'

    # Appends the string to the document
    doc.asis(xml_header)
    doc.asis(xml_schema)

    with tag('PO'):
        # Iterating rows for getting the values of each row
        # Use ws.max_row for all rows
        for row in ws.iter_rows(min_row=2, max_row=2, min_col=1, max_col=5):
            row = [cell.value for cell in row]
            wblist.append(row)
            with tag("PO_Information"):
                with tag("PO_Number"):
                    text(row[0])
                with tag("Site"):
                    text(row[1])
                with tag("PO_Description"):
                    text(row[2])
                with tag("Status"):
                    text(row[3])
                with tag("Revision"):
                    text(row[4])
        for row in ws.iter_rows(min_row=3, max_row= x, min_col=1, max_col=14):
            row = [cell.value for cell in row]
            wblist.append(row)
            with tag("PO_Line"):
                with tag("Line"):
                    text(row[5])
                with tag("Line_Type"):
                    text(row[6])
                with tag("Item"):
                    text(row[7])
                with tag("Item_Description"):
                    text(row[8])
                with tag("Quantity"):
                    text(row[9])
                with tag("Order_Unit"):
                    text(row[10])
                with tag("Manufacturer"):
                    text(row[11])
                with tag("Model_Number"):
                    text(row[12])
                with tag("GL_Account"):
                    text(row[13])

    result = indent(
        doc.getvalue(),
        indentation= '  ',
        indent_text=False
    )
    ##print(result)
    with open(f"fwn_{po}_{date}.xml", "w") as f:
        f.write(result)
    
clean_WB(w, x)
convert2xml('clean_wb.xlsx', x, date)
